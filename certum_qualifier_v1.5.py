#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════╗
║  CETRUM B2B — Cualificador Automático de Prospectos v1.5        ║
║  Capa 2 del Pipeline de Prospección                             ║
║                                                                 ║
║  Changelog v1.5 (auditoría senior):                             ║
║    [CRÍTICO] Session por hilo (thread-safe requests)            ║
║    [CRÍTICO] NO_WEB deduplicados por nombre en checkpoint       ║
║    [CRÍTICO] Guardia division-by-zero en estadísticas           ║
║    [ALTO]    Retry con backoff en fetch (1 reintento)           ║
║    [ALTO]    SSL verify=True primero, fallback verify=False     ║
║    [ALTO]    Límite de 5MB en respuesta HTTP                    ║
║    [ALTO]    Content-Type acepta text/html y xhtml              ║
║    [MEDIO]   soup.find("title") llamado una sola vez            ║
║    [BAJO]    Logging de retry y tamaño de respuesta             ║
║                                                                 ║
║  Conservado de v1.4:                                            ║
║    • Señales negativas activas en scoring                       ║
║    • Scoring por nombre en ERROR_FETCH y NO_WEB                 ║
║    • Checkpoint CSV con QUOTE_ALL (anti-corrupción)             ║
║    • Resume por website + nombre (para NO_WEB)                  ║
║    • Deduplicación por website en Excel final                   ║
║    • Email: regex estricta + priorización personal > genérico   ║
║    • Normalización de barrios de Madrid                         ║
║    • Flag ⚠ en emails con dominio cruzado                       ║
║    • Excel con pestañas Resumen/TierA/TierB/Todos               ║
║                                                                 ║
║  Uso:                                                           ║
║    pip install requests beautifulsoup4 openpyxl                  ║
║    python certum_qualifier_v1.5.py --input apify_raw.csv         ║
║                                                                 ║
║  Opciones:                                                      ║
║    --output   Excel de salida (default: prospectos_v1.5.xlsx)    ║
║    --delay    Segundos entre requests (default: 2.0)             ║
║    --timeout  Timeout HTTP (default: 12)                         ║
║    --workers  Hilos concurrentes (default: 2, no subir >5)       ║
║    --fresh    Ignorar checkpoint y reprocesar todo               ║
╚══════════════════════════════════════════════════════════════════╝
"""

import argparse
import csv
import logging
import os
import random
import re
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Optional
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ═══════════════════════════════════════════════════════════════
# CONFIGURACIÓN DE SCORING
# ═══════════════════════════════════════════════════════════════

SIGNALS_TIER_A = {
    r"\bcolombian[oa]s?\b": 15,
    r"\bcolombia\b": 12,
    r"\bapostilla[sr]?\b": 15,
    r"\bregistro[s]?\s+civil(?:es)?\b": 14,
    r"\bpartida\s+de\s+nacimiento\b": 12,
    r"\bcanciller[ií]a\b": 10,
    r"\bregistradur[ií]a\b": 10,
    r"\bdocumentos?\s+apostillados?\b": 12,
    r"\btr[aá]mites?\s+documentales?\b": 8,
    r"\bgesti[oó]n\s+documental\b": 8,
    r"\blegaliza(?:ci[oó]n|r)\b": 8,
}

SIGNALS_TIER_B = {
    r"\bextranjer[ií]a\b": 6,
    r"\binmigraci[oó]n\b": 6,
    r"\bnacionalidad\s+espa[nñ]ola\b": 8,
    r"\barraigo\b": 7,
    r"\breagrupaci[oó]n\s+familiar\b": 7,
    r"\bresidencia\b": 4,
    r"\bpermiso\s+de\s+trabajo\b": 5,
    r"\bvisado\b": 4,
    r"\bregularizaci[oó]n\b": 6,
    r"\bcarta\s+de\s+invitaci[oó]n\b": 5,
    r"\bnie\b": 3,
    r"\btie\b": 3,
    r"\bhomologaci[oó]n\b": 4,
    r"\biberoamericano[s]?\b": 6,
    r"\blatinoamericano[s]?\b": 5,
    r"\bmigrante[s]?\b": 4,
}

SIGNALS_NEGATIVE = {
    r"\bderecho\s+penal\b": -8,
    r"\baccidentes?\s+de\s+tr[aá]fico\b": -10,
    r"\bderecho\s+mercantil\b": -5,
    r"\bconcursal\b": -8,
    r"\bpropiedad\s+intelectual\b": -6,
    r"\bderecho\s+bancario\b": -6,
    r"\bdivorcio\b": -6,
}

THRESHOLD_A = 20
THRESHOLD_B = 8

GENERIC_EMAIL_PREFIXES = (
    "info@", "contacto@", "contact@", "admin@", "administracion@",
    "recepcion@", "hola@", "soporte@", "clientes@", "admon@",
)

EMAIL_BLACKLIST_FRAGMENTS = [
    ".png", ".jpg", ".gif", ".svg", ".css", ".js",
    "example.com", "email.com", "tu@", "your@", "usuario@dominio",
    "wixpress", "sentry", "webpack", "cloudflare",
    "tiktok.com", "facebook.com", "instagram.com",
    "//www.", "http", "filler@godaddy", "u003e",
    "lawfirm.com", "dominio.com",
]

MADRID_BARRIOS = {
    "salamanca", "centro", "retiro", "chamberí", "chamartín",
    "arganzuela", "moncloa", "latina", "carabanchel", "usera",
    "puente de vallecas", "moratalaz", "ciudad lineal", "hortaleza",
    "villaverde", "tetuán", "fuencarral", "barajas", "rivas-vaciamadrid",
}

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
]

EMAIL_STRICT_REGEX = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")

# Límite de respuesta HTTP (5MB) para evitar explotar RAM con páginas enormes
MAX_RESPONSE_BYTES = 5 * 1024 * 1024


# ═══════════════════════════════════════════════════════════════
# ESTRUCTURAS DE DATOS
# ═══════════════════════════════════════════════════════════════

@dataclass
class ProspectResult:
    nombre_despacho: str = ""
    website: str = ""
    email_principal: str = ""
    emails_encontrados: str = ""
    telefono: str = ""
    direccion: str = ""
    ciudad: str = ""
    rating_google: str = ""
    num_reviews: str = ""
    score: int = 0
    tier: str = "C"
    signals_matched: str = ""
    titulo_web: str = ""
    descripcion_web: str = ""
    tiene_formulario: bool = False
    estado_scrape: str = ""
    url_linkedin: str = ""


# ═══════════════════════════════════════════════════════════════
# LÓGICA CORE
# ═══════════════════════════════════════════════════════════════

# [FIX v1.5] Thread-local storage para sessions HTTP
_thread_local = threading.local()


def _get_session() -> requests.Session:
    """Retorna una Session exclusiva del hilo actual."""
    if not hasattr(_thread_local, "session"):
        _thread_local.session = requests.Session()
    return _thread_local.session


class CertumQualifier:
    def __init__(self, delay: float = 2.0, timeout: int = 12, workers: int = 2):
        self.delay = delay
        self.timeout = timeout
        self.workers = workers
        self.logger = logging.getLogger("CertumQualifier")

        self.patterns_a = {re.compile(p, re.IGNORECASE): w for p, w in SIGNALS_TIER_A.items()}
        self.patterns_b = {re.compile(p, re.IGNORECASE): w for p, w in SIGNALS_TIER_B.items()}
        self.patterns_neg = {re.compile(p, re.IGNORECASE): w for p, w in SIGNALS_NEGATIVE.items()}

        self.email_regex = re.compile(
            r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", re.IGNORECASE
        )
        self.linkedin_regex = re.compile(
            r"https?://(?:www\.)?linkedin\.com/(?:in|company)/[a-zA-Z0-9\-_%]+/?", re.IGNORECASE
        )

    def _get_headers(self) -> dict:
        return {
            "User-Agent": random.choice(USER_AGENTS),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
            "Connection": "close",
        }

    def _normalize_url(self, url: str) -> str:
        if not url:
            return ""
        url = url.strip().rstrip("/")
        if not url.startswith(("http://", "https://")):
            url = "https://" + url
        return url

    # [FIX v1.5] Fetch con retry, verify=True primero, límite de tamaño
    def _fetch_page(self, url: str) -> Optional[str]:
        session = _get_session()
        max_retries = 2

        for attempt in range(max_retries):
            # Primera iteración: verify=True. Segunda: verify=False (fallback SSL)
            verify = (attempt == 0)
            try:
                resp = session.get(
                    url,
                    headers=self._get_headers(),
                    timeout=(5, self.timeout),
                    allow_redirects=True,
                    verify=verify,
                    stream=True,
                )
                resp.raise_for_status()

                # [FIX v1.5] Validar Content-Type (html o xhtml)
                ct = resp.headers.get("Content-Type", "")
                if "text/html" not in ct and "application/xhtml" not in ct:
                    resp.close()
                    return None

                # [FIX v1.5] Límite de tamaño: leer max 5MB
                content_length = resp.headers.get("Content-Length")
                if content_length and int(content_length) > MAX_RESPONSE_BYTES:
                    self.logger.warning(f"  SKIP (>{MAX_RESPONSE_BYTES//1024//1024}MB): {url}")
                    resp.close()
                    return None

                chunks = []
                total = 0
                for chunk in resp.iter_content(chunk_size=65536, decode_unicode=False):
                    chunks.append(chunk)
                    total += len(chunk)
                    if total > MAX_RESPONSE_BYTES:
                        self.logger.warning(f"  TRUNCADO a {MAX_RESPONSE_BYTES//1024//1024}MB: {url}")
                        break
                resp.close()

                raw_bytes = b"".join(chunks)
                # Detectar encoding
                encoding = resp.encoding or resp.apparent_encoding or "utf-8"
                return raw_bytes.decode(encoding, errors="replace")

            except requests.exceptions.SSLError:
                if attempt == 0:
                    # Reintentar sin verificación SSL
                    continue
                return None
            except requests.exceptions.Timeout:
                if attempt == 0:
                    self.logger.debug(f"  Timeout (reintentando): {url}")
                    time.sleep(1)
                    continue
                return None
            except requests.exceptions.RequestException:
                if attempt == 0:
                    time.sleep(0.5)
                    continue
                return None
            except Exception:
                return None

        return None

    def _extract_text(self, html: str) -> tuple[str, str, str, bool, str]:
        soup = BeautifulSoup(html, "html.parser")
        for tag in soup(["script", "style", "noscript"]):
            tag.decompose()

        # [FIX v1.5] soup.find("title") una sola vez
        title_tag = soup.find("title")
        titulo = title_tag.get_text(strip=True) if title_tag else ""

        meta_desc = ""
        meta_tag = soup.find("meta", attrs={"name": "description"})
        if meta_tag:
            meta_desc = meta_tag.get("content", "")

        tiene_form = bool(soup.find("form"))

        linkedin_urls = set()
        for a_tag in soup.find_all("a", href=True):
            match = self.linkedin_regex.search(a_tag["href"])
            if match:
                linkedin_urls.add(match.group(0))

        text = re.sub(r"\s+", " ", soup.get_text(separator=" ", strip=True))
        return text, titulo, meta_desc, tiene_form, "; ".join(linkedin_urls)

    def _clean_email(self, email: str) -> Optional[str]:
        email = email.strip().lower()
        if not EMAIL_STRICT_REGEX.match(email):
            return None
        if any(frag in email for frag in EMAIL_BLACKLIST_FRAGMENTS):
            return None
        return email

    def _extract_emails_from_html(self, html: str) -> list[str]:
        raw = set(self.email_regex.findall(html))
        clean = set()
        for e in raw:
            validated = self._clean_email(e)
            if validated:
                clean.add(validated)
        return sorted(clean)

    def _score_text(self, text: str) -> tuple[int, int, int, list[str]]:
        """Retorna (score_total, score_a, score_b, signals)."""
        score_a, score_b, penalty, signals = 0, 0, 0, []

        for pattern, weight in self.patterns_a.items():
            matches = pattern.findall(text)
            if matches:
                pts = weight * min(len(matches), 3)
                score_a += pts
                signals.append(f"A:{pattern.pattern}(x{len(matches)}=+{pts})")

        for pattern, weight in self.patterns_b.items():
            matches = pattern.findall(text)
            if matches:
                pts = weight * min(len(matches), 3)
                score_b += pts
                signals.append(f"B:{pattern.pattern}(x{len(matches)}=+{pts})")

        for pattern, weight in self.patterns_neg.items():
            if pattern.search(text):
                penalty += abs(weight)
                signals.append(f"NEG:{pattern.pattern}({weight})")

        total = max(score_a + score_b - penalty, 0)
        return total, score_a, score_b, signals

    def _classify_tier(self, score_total: int, score_a: int) -> str:
        if score_total >= THRESHOLD_A and score_a >= 10:
            return "A"
        elif score_total >= THRESHOLD_B:
            return "B"
        return "C"

    def qualify_prospect(self, row: dict) -> ProspectResult:
        result = ProspectResult()

        result.nombre_despacho = (
            row.get("title") or row.get("name") or row.get("nombre") or ""
        ).strip()

        result.website = self._normalize_url(
            row.get("website") or row.get("site") or row.get("web") or ""
        )

        # Emails Apify con limpieza y priorización
        raw_apify = [
            str(v).strip().lower()
            for k, v in row.items()
            if str(k).startswith("emails/") and v and str(v).strip()
        ]
        apify_emails = sorted(set(e for e in raw_apify if self._clean_email(e)))
        personal = [e for e in apify_emails if not e.startswith(GENERIC_EMAIL_PREFIXES)]
        generic = [e for e in apify_emails if e.startswith(GENERIC_EMAIL_PREFIXES)]

        result.email_principal = (
            row.get("email")
            or (personal[0] if personal else "")
            or (generic[0] if generic else "")
        ).strip().lower()

        result.telefono = (row.get("phone") or row.get("telefono") or "").strip()
        result.direccion = (row.get("address") or row.get("full_address") or "").strip()

        result.ciudad = (row.get("city") or row.get("ciudad") or "").strip()
        if result.ciudad.lower() in MADRID_BARRIOS:
            result.ciudad = "Madrid"

        result.rating_google = str(row.get("rating") or row.get("google_rating") or "").strip()
        result.num_reviews = str(row.get("reviewsCount") or row.get("reviews") or "").strip()

        # Sin website → score por nombre
        if not result.website:
            result.estado_scrape = "NO_WEB"
            score, score_a, _, signals = self._score_text(result.nombre_despacho)
            result.score = score
            result.tier = self._classify_tier(score, score_a)
            result.signals_matched = "; ".join(signals) if signals else "Solo nombre"
            return result

        self.logger.info(f"  Analizando: {result.nombre_despacho[:40]} → {result.website}")

        html = self._fetch_page(result.website)

        # ERROR_FETCH → score por nombre
        if not html:
            result.estado_scrape = "ERROR_FETCH"
            score, score_a, _, signals = self._score_text(result.nombre_despacho)
            result.score = score
            result.tier = self._classify_tier(score, score_a)
            result.signals_matched = "; ".join(signals) if signals else "Solo nombre"
            return result

        text, titulo, meta_desc, tiene_form, linkedin = self._extract_text(html)
        result.titulo_web = titulo[:200]
        result.descripcion_web = meta_desc[:300]
        result.tiene_formulario = tiene_form
        result.url_linkedin = linkedin

        # Consolidar emails
        web_emails = self._extract_emails_from_html(html)
        all_emails = set(apify_emails) | set(web_emails)
        if result.email_principal:
            all_emails.add(result.email_principal)
        all_emails = {e for e in all_emails if self._clean_email(e)}
        result.emails_encontrados = "; ".join(sorted(all_emails))

        if not result.email_principal and web_emails:
            web_personal = [e for e in web_emails if not e.startswith(GENERIC_EMAIL_PREFIXES)]
            result.email_principal = web_personal[0] if web_personal else web_emails[0]

        # Scoring completo
        full_text = f"{result.nombre_despacho} {titulo} {meta_desc} {text}"
        score, score_a, score_b, signals = self._score_text(full_text)

        result.score = score
        result.tier = self._classify_tier(score, score_a)
        result.signals_matched = "; ".join(signals) if signals else "Sin señales"
        result.estado_scrape = "OK"
        return result


# ═══════════════════════════════════════════════════════════════
# I/O
# ═══════════════════════════════════════════════════════════════

def _read_input(path: str) -> list[dict]:
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        rows = []
        with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
            sample = f.read(4096)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            for row in csv.DictReader(f, dialect=dialect):
                rows.append(dict(row))
        return rows
    elif ext in (".xlsx", ".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(next(rows_iter))]
        rows = []
        for row_vals in rows_iter:
            rows.append({h: str(v) if v is not None else "" for h, v in zip(headers, row_vals)})
        wb.close()
        return rows
    else:
        logging.error(f"Formato no soportado: {ext}. Usa .csv o .xlsx")
        sys.exit(1)


def _safe_pct(numerator: int, denominator: int) -> str:
    """[FIX v1.5] División segura para porcentajes."""
    if denominator == 0:
        return "0.0%"
    return f"{numerator / denominator * 100:.1f}%"


def _write_excel(results: list[ProspectResult], path: str):
    if not HAS_OPENPYXL:
        csv_path = path.replace(".xlsx", ".csv")
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            if results:
                writer = csv.DictWriter(f, fieldnames=asdict(results[0]).keys())
                writer.writeheader()
                for r in results:
                    writer.writerow(asdict(r))
        logging.info(f"openpyxl no disponible. CSV en: {csv_path}")
        return

    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    fill_a = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    fill_b = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    fill_all = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    fill_sum = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='BDBDBD'),
        right=Side(style='thin', color='BDBDBD'),
        top=Side(style='thin', color='BDBDBD'),
        bottom=Side(style='thin', color='BDBDBD'),
    )

    COLS = [
        "Tier", "Score", "Despacho", "Email Principal", "Otros Emails",
        "Teléfono", "Website", "Ciudad", "Dirección", "Rating", "Reseñas",
        "Título Web", "Descripción Web", "LinkedIn", "Formulario",
        "Señales Detectadas", "Estado Scrape",
    ]
    WIDTHS = [5, 6, 40, 35, 35, 18, 40, 12, 30, 6, 8, 40, 50, 30, 10, 50, 15]

    def write_sheet(ws, records, hdr_fill):
        ws.append(COLS)
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for r in sorted(records, key=lambda x: x.score, reverse=True):
            ws.append([
                r.tier, r.score, r.nombre_despacho, r.email_principal,
                r.emails_encontrados, r.telefono, r.website, r.ciudad,
                r.direccion, r.rating_google, r.num_reviews,
                r.titulo_web, r.descripcion_web, r.url_linkedin,
                "Sí" if r.tiene_formulario else "No",
                r.signals_matched, r.estado_scrape,
            ])
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.font = Font(name="Arial", size=9)
            if "⚠" in str(r.email_principal):
                ws.cell(row=ws.max_row, column=4).fill = warn_fill
        for i, w in enumerate(WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    tier_a = [r for r in results if r.tier == "A"]
    tier_b = [r for r in results if r.tier == "B"]
    tier_c = [r for r in results if r.tier == "C"]
    email_ab = sum(1 for r in tier_a + tier_b if r.email_principal and "⚠" not in r.email_principal)

    ws_s = wb.active
    ws_s.title = "Resumen"
    for row in [
        ["CETRUM B2B — Cualificación v1.5"],
        [],
        ["Métrica", "Valor"],
        ["Total prospectos (deduplicados)", len(results)],
        ["TIER A — Alta prioridad", len(tier_a)],
        ["TIER B — Extranjería genérica", len(tier_b)],
        ["TIER C — Irrelevante / sin datos", len(tier_c)],
        ["Tasa cualificación (A+B)", _safe_pct(len(tier_a) + len(tier_b), len(results))],
        [],
        ["Email válido (Tier A+B)", email_ab],
        ["Scrape OK", sum(1 for r in results if r.estado_scrape == "OK")],
        ["ERROR_FETCH", sum(1 for r in results if r.estado_scrape == "ERROR_FETCH")],
    ]:
        ws_s.append(row)
    ws_s["A1"].font = Font(bold=True, name="Arial", size=14, color="1A237E")
    for cell in ws_s[3]:
        cell.font = hdr_font
        cell.fill = fill_sum
    ws_s.column_dimensions["A"].width = 40
    ws_s.column_dimensions["B"].width = 20

    if tier_a:
        write_sheet(wb.create_sheet("TIER A — Prioridad Alta"), tier_a, fill_a)
    if tier_b:
        write_sheet(wb.create_sheet("TIER B — Extranjería"), tier_b, fill_b)
    write_sheet(wb.create_sheet("Todos los Prospectos"), results, fill_all)
    wb.save(path)


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="CETRUM Qualifier v1.5")
    parser.add_argument("--input", required=True, help="Archivo CSV/XLSX de entrada")
    parser.add_argument("--output", default="prospectos_v1.5.xlsx", help="Excel de salida")
    parser.add_argument("--workers", type=int, default=2, help="Hilos concurrentes")
    parser.add_argument("--delay", type=float, default=2.0, help="Delay entre requests")
    parser.add_argument("--timeout", type=int, default=12, help="Timeout HTTP")
    parser.add_argument("--fresh", action="store_true", help="Ignorar checkpoint")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: El archivo {args.input} no existe.")
        sys.exit(1)

    log_file = "qualifier_v1.5_log.txt"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s │ %(message)s",
        datefmt="%H:%M:%S",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )
    logger = logging.getLogger()

    checkpoint_file = "checkpoint_v1.5.csv"
    CHECKPOINT_FIELDS = list(asdict(ProspectResult()).keys())

    # [FIX v1.5] Resume por website Y por nombre (para cubrir NO_WEB)
    processed_websites = set()
    processed_names = set()

    if not args.fresh and os.path.exists(checkpoint_file):
        try:
            with open(checkpoint_file, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    w = (row.get("website") or "").strip().lower().rstrip("/")
                    n = (row.get("nombre_despacho") or "").strip().lower()
                    if w:
                        processed_websites.add(w)
                    if n:
                        processed_names.add(n)
            logger.info(f"♻ Resumiendo: {len(processed_websites)} websites + {len(processed_names)} nombres ya procesados.")
        except Exception as e:
            logger.warning(f"⚠ Error leyendo checkpoint ({e}), iniciando desde cero.")
    elif args.fresh and os.path.exists(checkpoint_file):
        os.remove(checkpoint_file)
        logger.info("🗑 Checkpoint eliminado (--fresh).")

    rows = _read_input(args.input)
    if not rows:
        logger.error("Sin filas en el archivo de entrada.")
        sys.exit(1)

    # [FIX v1.5] Deduplicación en resume: usa website para los que tienen, nombre para NO_WEB
    to_process = []
    for row in rows:
        url = (row.get("website") or row.get("site") or row.get("web") or "").strip().lower().rstrip("/")
        name = (row.get("title") or row.get("name") or "").strip().lower()
        if url:
            if url not in processed_websites:
                to_process.append(row)
        else:
            if name not in processed_names:
                to_process.append(row)

    logger.info(f"Archivo: {len(rows)} | Procesados: {len(processed_websites)} + {len(processed_names)} sin web | Pendientes: {len(to_process)}")

    qualifier = CertumQualifier(delay=args.delay, timeout=args.timeout, workers=args.workers)

    file_is_new = not os.path.exists(checkpoint_file)
    checkpoint_fh = open(checkpoint_file, "a", encoding="utf-8-sig", newline="")
    writer = csv.DictWriter(checkpoint_fh, fieldnames=CHECKPOINT_FIELDS, quoting=csv.QUOTE_ALL)
    if file_is_new:
        writer.writeheader()
        checkpoint_fh.flush()

    done_count = 0
    total_target = len(to_process) + len(processed_websites) + len(processed_names)

    logger.info(f"Workers: {args.workers} | Delay: {args.delay}s | Timeout: {args.timeout}s")
    logger.info("=" * 60)

    def process_one(row):
        time.sleep(args.delay * random.uniform(0.5, 1.5))
        return qualifier.qualify_prospect(row)

    try:
        with ThreadPoolExecutor(max_workers=args.workers) as executor:
            future_to_row = {executor.submit(process_one, r): r for r in to_process}
            for future in as_completed(future_to_row):
                done_count += 1
                try:
                    res = future.result()
                    writer.writerow(asdict(res))
                    checkpoint_fh.flush()
                    progress = done_count + len(processed_websites) + len(processed_names)
                    logger.info(
                        f"  [{progress}/{total_target}] "
                        f"{res.nombre_despacho[:35]:35s} │ T-{res.tier} │ s={res.score:>3} │ {res.estado_scrape}"
                    )
                except Exception as e:
                    logger.error(f"  Error hilo: {e}")
    except KeyboardInterrupt:
        logger.warning("\n⚠ Interrumpido. Progreso guardado en checkpoint.")
    finally:
        checkpoint_fh.close()

    # ── Consolidación final ──
    logger.info("=" * 60)
    logger.info("Consolidando...")

    all_results = []
    seen_websites = set()
    seen_names_no_web = set()
    with open(checkpoint_file, "r", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            w = (row.get("website") or "").strip().lower().rstrip("/")
            n = (row.get("nombre_despacho") or "").strip().lower()

            # Deduplicar: por website si tiene, por nombre si NO_WEB
            if w:
                if w in seen_websites:
                    continue
                seen_websites.add(w)
            else:
                if n in seen_names_no_web:
                    continue
                seen_names_no_web.add(n)

            pr = ProspectResult(
                nombre_despacho=row.get("nombre_despacho", ""),
                website=row.get("website", ""),
                email_principal=row.get("email_principal", ""),
                emails_encontrados=row.get("emails_encontrados", ""),
                telefono=row.get("telefono", ""),
                direccion=row.get("direccion", ""),
                ciudad=row.get("ciudad", ""),
                rating_google=row.get("rating_google", ""),
                num_reviews=row.get("num_reviews", ""),
                score=int(row.get("score") or 0),
                tier=row.get("tier", "C"),
                signals_matched=row.get("signals_matched", ""),
                titulo_web=row.get("titulo_web", ""),
                descripcion_web=row.get("descripcion_web", ""),
                tiene_formulario=row.get("tiene_formulario", "") == "True",
                estado_scrape=row.get("estado_scrape", ""),
                url_linkedin=row.get("url_linkedin", ""),
            )
            all_results.append(pr)

    # Flag emails con dominio cruzado
    for r in all_results:
        email = r.email_principal.strip().lower()
        website = r.website.strip().lower()
        if email and "@" in email and website:
            email_domain = email.split("@")[1]
            try:
                web_domain = urlparse(website).netloc.replace("www.", "")
            except Exception:
                continue
            FREE = ("gmail.com", "hotmail.com", "hotmail.es", "yahoo.com", "yahoo.es",
                    "outlook.com", "outlook.es", "ymail.com", "live.com",
                    "icam.es", "icab.cat", "icav.es", "icali.es")
            if (email_domain != web_domain
                    and email_domain not in web_domain
                    and web_domain not in email_domain
                    and email_domain not in FREE):
                email_root = re.sub(r"\.(com|es|org|net|eu|info)$", "", email_domain)
                web_root = re.sub(r"\.(com|es|org|net|eu|info)$", "", web_domain)
                if email_root != web_root:
                    r.email_principal = f"⚠ REVISAR: {email}"

    tier_a = sum(1 for r in all_results if r.tier == "A")
    tier_b = sum(1 for r in all_results if r.tier == "B")
    tier_c = sum(1 for r in all_results if r.tier == "C")

    logger.info(f"  Total: {len(all_results)} | A: {tier_a} | B: {tier_b} | C: {tier_c}")
    logger.info(f"  Tasa A+B: {_safe_pct(tier_a + tier_b, len(all_results))}")

    _write_excel(all_results, args.output)
    logger.info(f"\n✅ {args.output}")


if __name__ == "__main__":
    main()
